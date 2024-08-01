import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import math
import hashlib
import sqlite3
import openpyxl
import secrets
import smtplib
from email.mime.text import MIMEText

# Create a SQLite database connection
conn = sqlite3.connect('finfusion.db')
c = conn.cursor()

# Create tables if they don't exist
c.execute('''CREATE TABLE IF NOT EXISTS users
             (username TEXT PRIMARY KEY, password TEXT)''')
c.execute('''CREATE TABLE IF NOT EXISTS financial_data
             (username TEXT, date DATE, description TEXT, amount REAL, type TEXT)''')
c.execute('''CREATE TABLE IF NOT EXISTS password_recovery
             (username TEXT, token TEXT)''')
conn.commit()

# Function to hash passwords
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# Function to check if user exists
def user_exists(username):
    c.execute("SELECT 1 FROM users WHERE username=?", (username,))
    return c.fetchone() is not None

# Function to authenticate user
def authenticate_user(username, password):
    c.execute("SELECT 1 FROM users WHERE username=? AND password=?", (username, hash_password(password)))
    return c.fetchone() is not None

# Function to add new user
def add_user(username, password):
    c.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, hash_password(password)))
    conn.commit()

# Function to add financial data
def add_financial_data(username, date, description, amount, type):
    c.execute("INSERT INTO financial_data (username, date, description, amount, type) VALUES (?, ?, ?, ?, ?)", (username, date, description, amount, type))
    conn.commit()

# Function to generate a random token
def generate_token():
    return secrets.token_urlsafe(16)

# Function to send a recovery email
def send_recovery_email(username, token):
    # Configuração do servidor de e-mail
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login('seu_email@gmail.com', 'sua_senha')
    
    # Mensagem de e-mail
    msg = MIMEText(f'Clique no link para redefinir sua senha: http://localhost:8501/reset-password/{token}')
    msg['Subject'] = 'Redefinição de Senha'
    msg['From'] = 'seu_email@gmail.com'
    msg['To'] = username
    
    # Enviar e-mail
    server.sendmail('seu_email@gmail.com', username, msg.as_string())
    server.quit()

# Função para solicitar recuperação de senha
def request_password_recovery(username):
    # Verificar se o usuário existe
    if user_exists(username):
        # Gerar token e armazenar na tabela
        token = generate_token()
        c.execute("INSERT INTO password_recovery (username, token) VALUES (?, ?)", (username, token))
        conn.commit()
        
        # Enviar e-mail com o token
        send_recovery_email(username, token)
        st.success('E-mail de recuperação enviado com sucesso!')
    else:
        st.error('Usuário não encontrado')

# Função para exportar dados financeiros para Excel
def export_to_excel(financial_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Cabeçalho
    ws['A1'] = 'Data'
    ws['B1'] = 'Descrição'
    ws['C1'] = 'Quantia'
    ws['D1'] = 'Tipo'
    
    # Dados
    for i, row in enumerate(financial_data):
        ws[f'A{i+2}'] = row[1]
        ws[f'B{i+2}'] = row[2]
        ws[f'C{i+2}'] = row[3]
        ws[f'D{i+2}'] = row[4]
    
    # Salvar arquivo
    wb.save('financial_data.xlsx')

# Função para importar dados financeiros de Excel
def import_from_excel(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    
    # Dados
    financial_data = []
    for row in ws.iter_rows(min_row=2):
        financial_data.append((row[0].value, row[1].value, row[2].value, row[3].value))
    
    return financial_data

# Função para obter dados financeiros do usuário
def get_financial_data(username, filter=None):
    if filter:
        c.execute("SELECT * FROM financial_data WHERE username=? AND type=?", (username, filter))
    else:
        c.execute("SELECT * FROM financial_data WHERE username=?", (username,))
    return c.fetchall()

# Login form